from openpyxl import Workbook, load_workbook
from ostertage import ostern, easter_day


wb = load_workbook("Ostersonntag.xlsx")
ws = wb["Ostersonntag"]
year = ws["A1"].value

easter = ostern(year)
easter_days = easter_day(easter)

row = 2
for day in easter_days:
    row += 1
    ws[f"A{row}"] = day[0]
    ws[f"B{row}"] = day[1]
    print(f"{day[0]}: {day[1]:%d.%m.%Y}")

wb.save("Ostersonntag.xlsx")

