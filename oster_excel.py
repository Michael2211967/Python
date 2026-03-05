#!/usr/bin/python3
from openpyxl import Workbook, load_workbook
from Ostersonntag import ostern

wb = load_workbook("Ostersonntag.xlsx")
ws = wb["Ostersonntag"]

year = int(input("Geben Sie bitte das gewünschte Jahr ein: "))

ws["A1"].value = year

for interval_year in range(11):
    easter = ostern(year + interval_year)

    print(f"{easter:%d.%m.%Y}")
    ws[f"A{3 + interval_year}"].value = easter

wb.save("Ostersonntag.xlsx")
