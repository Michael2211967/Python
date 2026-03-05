from dateutil.easter import *
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook

year = int(input("Für welches Jahr wollen Sie die Ostertage ermitteln? "))

try:
    workbook = load_workbook("easter.xlsx")
except:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ostern"
    
sheet = workbook["Ostern"]
sheet["A1"].value = year

# Berechne das Datum von Ostern im westlichen Kalender für das Jahr 2023
easter_date = easter(year)

# Gib das Datum von Ostern aus
print("Das Datum von Ostersonntag im Jahr {} ist: {}".format(year, easter_date.strftime("%d.%m.%Y")))

# Weiberfastnacht berechnen (52 Tage vor Ostern)
weiberfastnacht = easter_date - timedelta(days=52)
sheet["A3"].value = "Weiberfastnacht"
sheet["B3"].value = weiberfastnacht.strftime("%d.%m.%Y")

# Rosenmontag berechnen (48 Tage vor Ostern)
rosenmontag = easter_date - timedelta(days=48)
sheet["A4"].value = "Rosenmontag"
sheet["B4"].value = rosenmontag.strftime("%d.%m.%Y")

# Aschermittwoch berechnen (46 Tage vor Ostern)
aschermittwoch = easter_date - timedelta(days=46)
sheet["A5"].value = "Aschermittwoch"
sheet["B5"].value = aschermittwoch.strftime("%d.%m.%Y")

# Karfreitag berechnen (2 Tage vor Ostern)
karfreitag = easter_date - timedelta(days=2)
sheet["A6"].value = "Karfreitag"
sheet["B6"].value = karfreitag.strftime("%d.%m.%Y")
sheet["A7"].value = "Ostersonntag"
sheet["B7"].value = easter_date.strftime("%d.%m.%Y")

# Ostermontag berechnen (1 Tag nach Ostern)
ostermontag = easter_date + timedelta(days=1)
sheet["A8"].value = "Ostermontag"
sheet["B8"].value = ostermontag.strftime("%d.%m.%Y")

# Himmelfahrt berechnen (39 Tage nach Ostern)
himmelfahrt = easter_date + timedelta(days=39)
sheet["A9"].value = "Christi Himmelfahrt"
sheet["B9"].value = himmelfahrt.strftime("%d.%m.%Y")

# Pfingstsonntag berechnen (49 Tage nach Ostern)
pfingstsonntag = easter_date + timedelta(days=49)
sheet["A10"].value = "Pfingstsonntag"
sheet["B10"].value = pfingstsonntag.strftime("%d.%m.%Y")

# Pfingstmontag berechnen (50 Tage nach Ostern)
pfingstmontag = easter_date + timedelta(days=50)
sheet["A11"].value = "Pfingstmontag"
sheet["B11"].value = pfingstmontag.strftime("%d.%m.%Y")

# Fronleichnam berechnen (60 Tage nach Ostern)
fronleichnam = easter_date + timedelta(days=60)
sheet["A12"].value = "Fronleichnam"
sheet["B12"].value = fronleichnam.strftime("%d.%m.%Y")

workbook.save("easter.xlsx")

# Ergebnisse ausgeben
print("Weiberfastnacht: ", weiberfastnacht.strftime("%d.%m.%Y"))
print("Rosenmontag: ", rosenmontag.strftime("%d.%m.%Y"))
print("Aschermittwoch: ", aschermittwoch.strftime("%d.%m.%Y"))
print("Karfreitag: ", karfreitag.strftime("%d.%m.%Y"))
print("Ostermontag: ", ostermontag.strftime("%d.%m.%Y"))
print("Himmelfahrt: ", himmelfahrt.strftime("%d.%m.%Y"))
print("Pfingstsonntag: ", pfingstsonntag.strftime("%d.%m.%Y"))
print("Pfingstmontag: ", pfingstmontag.strftime("%d.%m.%Y"))
print("Fronleichnam: ", fronleichnam.strftime("%d.%m.%Y"))
